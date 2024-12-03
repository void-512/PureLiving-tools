import re
import sys
import configparser
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

cfg = 'FapiaoUpdaterConfig.cfg'

configObj = configparser.ConfigParser()
configObj.read(cfg)
pivot = configObj.get('src', 'pivot')
src = configObj.get('src', 'src')
src_sheet_name = configObj.get('src', 'sheet')

src_workbook = load_workbook(src)
dest_workbook = load_workbook('Content Updated.xlsx')
src_sheet = src_workbook[src_sheet_name]
dest_sheet = dest_workbook.active

def separate_row_col(cell_loc):
    match = re.match(r"([A-Z]+)(\d+)", cell_loc)
    if match:
        return match.group(1), match.group(2)
    else:
        sys.exit('Error occurs, cannot separate COL id and ROW id')

def address_translator(src_sheet, dest_sheet, old_cell):
    global pivot
    col_id = old_cell.column
    row_id = old_cell.row
    old_cell_col_title = src_sheet.cell(column=col_id, row=1).value

    # Find the pivot column in the source sheet
    pivot_col_src = None
    for cell in src_sheet[1]:  # Access the first row
        if cell.value == pivot:
            pivot_col_src = cell.column
            break
    if pivot_col_src is None:
        raise ValueError(f"Pivot '{pivot}' not found in source sheet")

    # Find the value at the pivot column and old cell's row
    pivot_val = src_sheet.cell(column=pivot_col_src, row=row_id).value

    # Find the pivot column in the destination sheet
    pivot_col_dest = None
    for cell in dest_sheet[1]:  # Access the first row
        if cell.value == pivot:
            pivot_col_dest = cell.column
            break
    if pivot_col_dest is None:
        raise ValueError(f"Pivot '{pivot}' not found in destination sheet")

    # Find the column in the destination sheet corresponding to the old cell's column title
    new_cell_col_id = None
    for cell in dest_sheet[1]:  # Access the first row
        if cell.value == old_cell_col_title:
            new_cell_col_id = cell.column
            break
    if new_cell_col_id is None:
        return None

    # Find the row in the destination sheet where the pivot value matches
    new_cell_row_id = None
    for row in dest_sheet.iter_rows(min_row=2, min_col=pivot_col_dest, max_col=pivot_col_dest, values_only=False):
        for cell in row:
            if cell.value == pivot_val:
                new_cell_row_id = cell.row
                break
        if new_cell_row_id:
            break
    if new_cell_row_id is None:
        return None

    # Return the translated cell
    return dest_sheet.cell(column=new_cell_col_id, row=new_cell_row_id)



class CellColor:
    def __init__(self, color, col, row):
        self.color = color
        self.col = col
        self.row = row

# Iterate over all columns
def construct_color_list(src_sheet, dest_sheet):
    color_not_empty = []
    for col in src_sheet.iter_cols(min_row=1, max_row=src_sheet.max_row, min_col=1, max_col=src_sheet.max_column, values_only=False):
        for cell in col:
            color_code = cell.fill.start_color.index
            if color_code is None or color_code == '00000000':
                continue
            else:
                target_cell = address_translator(src_sheet, dest_sheet, cell)
                if target_cell is None:
                    continue
                color_not_empty.append(CellColor(color_code, target_cell.column, target_cell.row))
    return color_not_empty

newlist = construct_color_list(src_sheet, dest_sheet)

for item in newlist:
    cell = dest_sheet.cell(column=item.col, row=item.row)
    fill = PatternFill(start_color=item.color, 
                   end_color=item.color,    
                   fill_type="solid")
    cell.fill = fill

dest_workbook.save('Highlight Synchronized.xlsx')